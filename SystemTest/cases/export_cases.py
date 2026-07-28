# -*- coding: utf-8 -*-
"""Export system test cases from docs/系统测试用例.xlsx into case_index.json."""
from __future__ import annotations

import json
from pathlib import Path

from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parents[2]
XLSX = ROOT / "docs" / "系统测试用例.xlsx"
OUT = Path(__file__).resolve().parent / "case_index.json"


def main() -> None:
    wb = load_workbook(XLSX, data_only=True)
    ws = wb["Test Cases"]
    cases = []
    for row in range(5, ws.max_row + 1):
        tid = ws.cell(row, 1).value
        if not tid:
            continue
        cases.append(
            {
                "id": str(tid).strip(),
                "item": ws.cell(row, 2).value,
                "type": ws.cell(row, 3).value,
                "title": ws.cell(row, 4).value,
                "criticality": ws.cell(row, 5).value,
                "precondition": ws.cell(row, 6).value,
                "procedure": ws.cell(row, 7).value,
                "expected": ws.cell(row, 8).value,
                "remark": ws.cell(row, 13).value,
            }
        )
    OUT.write_text(json.dumps(cases, ensure_ascii=False, indent=2), encoding="utf-8")
    print(f"wrote {len(cases)} cases -> {OUT}")


if __name__ == "__main__":
    main()
