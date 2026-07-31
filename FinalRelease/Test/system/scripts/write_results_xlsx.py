# -*- coding: utf-8 -*-
"""Write API system-test results into results/系统测试执行结果-*.xlsx from junit XML."""
from __future__ import annotations

import argparse
import re
import xml.etree.ElementTree as ET
from datetime import datetime
from pathlib import Path
from shutil import copy2

from openpyxl import load_workbook

FINAL_RELEASE = Path(__file__).resolve().parents[3]
REPO_ROOT = FINAL_RELEASE.parent
SYSTEM = Path(__file__).resolve().parents[1]
SRC_XLSX = FINAL_RELEASE / "docs" / "系统测试用例.xlsx"
RESULTS = SYSTEM / "results"
TC_RE = re.compile(r"TC[_-]?(\d{3})", re.I)


def parse_junit(path: Path) -> dict[str, str]:
    """Return {TC-001: Y/N} from pytest junit classname/name/properties."""
    if not path.exists():
        return {}
    root = ET.parse(path).getroot()
    outcomes: dict[str, str] = {}
    for case in root.iter("testcase"):
        blob = f"{case.get('classname', '')} {case.get('name', '')}"
        # also search markers in name like tc_001
        ids = set()
        for match in TC_RE.finditer(blob):
            ids.add(f"TC-{match.group(1)}")
        # pytest mark may appear in properties
        for prop in case.findall("properties/property"):
            if prop.get("name") == "tc" and prop.get("value"):
                ids.add(prop.get("value", "").strip().upper().replace("_", "-"))
        failed = case.find("failure") is not None or case.find("error") is not None
        skipped = case.find("skipped") is not None
        for tid in ids:
            if skipped:
                continue
            outcomes[tid] = "N" if failed else "Y"
    return outcomes


def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument("--junit", type=Path, required=True)
    parser.add_argument("--stamp", default=datetime.now().strftime("%Y%m%d-%H%M%S"))
    args = parser.parse_args()

    RESULTS.mkdir(parents=True, exist_ok=True)
    outcomes = parse_junit(args.junit)

    summary = RESULTS / f"api-summary-{args.stamp}.md"
    lines = [
        f"# API 系统测试摘要 ({args.stamp})",
        "",
        f"- junit: `{args.junit.name}`",
        f"- 回填用例数: {len(outcomes)}",
        f"- 通过: {sum(1 for v in outcomes.values() if v == 'Y')}",
        f"- 失败: {sum(1 for v in outcomes.values() if v == 'N')}",
        "",
        "| TC | Status |",
        "|----|--------|",
    ]
    for tid in sorted(outcomes):
        lines.append(f"| {tid} | {outcomes[tid]} |")
    summary.write_text("\n".join(lines) + "\n", encoding="utf-8")
    print(f"wrote {summary}")

    # Prefer ASCII filename to avoid editor locks on CJK paths; fall back to temp then replace.
    out = RESULTS / f"system-test-results-{args.stamp}.xlsx"
    tmp = RESULTS / f"_tmp-results-{args.stamp}.xlsx"
    copy2(SRC_XLSX, tmp)
    wb = load_workbook(tmp)
    ws = wb["Test Cases"]
    filled = 0
    for row in range(5, ws.max_row + 1):
        tid = ws.cell(row, 1).value
        if not tid:
            continue
        tid = str(tid).strip()
        if tid not in outcomes:
            continue
        ws.cell(row, 10, outcomes[tid])  # Status 是否通过
        ws.cell(row, 9, "API 自动化" if outcomes[tid] == "Y" else "API 自动化失败")
        remark = ws.cell(row, 13).value or ""
        ws.cell(row, 13, f"{remark}; auto@{args.stamp}".strip("; "))
        filled += 1
    try:
        wb.save(tmp)
        tmp.replace(out)
        print(f"wrote {out} (filled {filled})")
    except PermissionError as exc:
        alt = RESULTS / f"system-test-results-{args.stamp}-alt.xlsx"
        wb.save(alt)
        print(f"permission denied for {out}; wrote {alt} instead ({exc})")
    finally:
        if tmp.exists() and tmp != out:
            try:
                tmp.unlink()
            except OSError:
                pass


if __name__ == "__main__":
    main()
